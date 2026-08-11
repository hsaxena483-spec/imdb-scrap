import os
import random
import time
import json
import sys
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from dotenv import load_dotenv
import openpyxl

import db

# Load environment variables relative to scraper.py directory
basedir = os.path.abspath(os.path.dirname(__file__))
load_dotenv(os.path.join(basedir, '.env'))

SCRAPE_LIMIT = int(os.getenv("SCRAPE_LIMIT", "10"))
REVIEWS_LIMIT = int(os.getenv("REVIEWS_LIMIT", "10"))
RANDOM_DELAY_MIN = float(os.getenv("RANDOM_DELAY_MIN", "2.0"))
RANDOM_DELAY_MAX = float(os.getenv("RANDOM_DELAY_MAX", "5.0"))
EXCEL_FILE = os.getenv("EXCEL_FILE", "C:\\Users\\saxen\\Downloads\\imdb analysis.xlsx")

sys.stdout.reconfigure(encoding='utf-8')

def init_driver():
    """
    Initializes headless Chrome webdriver with optimized arguments (eager load, block images).
    """
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--log-level=3")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    # Enable eager page load strategy (stops waiting for images and subresources)
    options.page_load_strategy = 'eager'
    
    # Block images and flash content to save bandwidth and load faster
    prefs = {
        "profile.managed_default_content_settings.images": 2,
        "profile.default_content_setting_values.notifications": 2
    }
    options.add_experimental_option("prefs", prefs)
    
    # Suppress unnecessary messages and disable automation indicators
    options.add_experimental_option('excludeSwitches', ['enable-logging', 'enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)
    
    # Anti-bot bypass (removes navigator.webdriver indicator)
    options.add_argument("--disable-blink-features=AutomationControlled")
    
    # Detect Chrome binary path (useful for Linux containers and Render)
    chrome_bin = os.getenv("CHROME_BIN")
    if chrome_bin:
        options.binary_location = chrome_bin
    elif os.path.exists("/usr/bin/google-chrome-stable"):
        options.binary_location = "/usr/bin/google-chrome-stable"
    elif os.path.exists("/usr/bin/google-chrome"):
        options.binary_location = "/usr/bin/google-chrome"
        
    try:
        # In modern Selenium (>= 4.6.0), Selenium Manager automatically downloads/locates Chrome & chromedriver.
        # This is the cleanest way and works out-of-the-box when Chrome is installed on the system path.
        driver = webdriver.Chrome(options=options)
    except Exception as e:
        print(f"Direct Chrome initialization failed ({e}). Falling back to ChromeDriverManager...")
        try:
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        except Exception as fallback_err:
            print(f"Fallback to ChromeDriverManager also failed: {fallback_err}")
            raise fallback_err
            
    try:
        driver.set_page_load_timeout(30)
        driver.set_script_timeout(30)
    except Exception as timeout_err:
        print(f"Warning: Failed to set driver timeouts: {timeout_err}")
        
    return driver

def get_next_data(driver, url):
    """
    Loads page, waits randomly to mimic human activity, and extracts
    the __NEXT_DATA__ JSON script payload.
    """
    delay = random.uniform(RANDOM_DELAY_MIN, RANDOM_DELAY_MAX)
    print(f"Loading URL: {url} (waiting {delay:.2f}s before fetching...)")
    time.sleep(delay)
    
    try:
        driver.get(url)
        # Wait to ensure page navigation, Javascript rendering, and WAF challenge reloads complete
        time.sleep(4.0)
        
        soup = BeautifulSoup(driver.page_source, "lxml")
        next_data_script = soup.find("script", id="__NEXT_DATA__")
        
        if next_data_script and next_data_script.string:
            data = json.loads(next_data_script.string)
            try:
                # Add og:image as fallback poster url in data
                og_image = soup.find('meta', property='og:image')
                if og_image and og_image.get('content') and not 'social/imdb_logo' in og_image.get('content', ''):
                    data['fallback_poster_url'] = og_image['content']
            except Exception as meta_err:
                pass
            return data
        else:
            print(f"Warning: No __NEXT_DATA__ script tag found for {url}")
            return None
    except (TimeoutException, WebDriverException) as e:
        raise e
    except Exception as e:
        print(f"Error fetching page {url}: {e}")
        return None

def parse_release_date(date_dict):
    """
    Takes release date dict {day, month, year} and formats it as YYYY-MM-DD.
    """
    if not date_dict or not isinstance(date_dict, dict):
        return None
    year = date_dict.get('year')
    month = date_dict.get('month') or 1
    day = date_dict.get('day') or 1
    if year:
        return f"{year:04d}-{month:02d}-{day:02d}"
    return None

def extract_tv_shows_list(next_data):
    """
    Parses the tvmeter chart JSON payload to extract show details.
    """
    shows = []
    if not next_data:
        return shows
        
    edges = next_data.get("props", {}).get("pageProps", {}).get("pageData", {}).get("chartTitles", {}).get("edges", [])
    print(f"Total shows in chart data: {len(edges)}")
    
    for idx, edge in enumerate(edges):
        node = edge.get("node", {})
        if not node:
            continue
            
        show_id = node.get("id")
        title = node.get("titleText", {}).get("text", "Unknown Title")
        show_type = node.get("titleType", {}).get("id")
        
        release_year = node.get("releaseYear", {}).get("year")
        end_year = node.get("releaseYear", {}).get("endYear")
        
        ratings = node.get("ratingsSummary", {})
        global_rating = ratings.get("aggregateRating")
        global_vote_count = ratings.get("voteCount")
        
        runtime = node.get("runtime", {})
        runtime_seconds = runtime.get("seconds") if runtime else None
        
        cert = node.get("certificate", {})
        certificate = cert.get("rating") if cert else None
        
        plot_obj = node.get("plot", {})
        plot = plot_obj.get("plotText", {}).get("plainText") if plot_obj else None
        
        img = node.get("primaryImage", {})
        poster_url = img.get("url") if img else None
        
        release_date = parse_release_date(node.get("releaseDate"))
        
        episodes_obj = node.get("episodes", {})
        total_episodes = None
        if episodes_obj and episodes_obj.get("episodes"):
            total_episodes = episodes_obj.get("episodes", {}).get("total")
            
        # Parse creators and stars
        creators_list = []
        stars_list = []
        for credit in node.get("principalCreditsV2", []):
            grouping = credit.get("grouping", {})
            group_text = grouping.get("text", "").lower()
            group_id = grouping.get("groupingId", "").lower()
            
            is_creator = "creator" in group_text or "creator" in group_id or "writer" in group_text or "director" in group_text
            is_star = "star" in group_text or "star" in group_id or "actor" in group_text or "cast" in group_text
            
            names = []
            for cred in credit.get("credits", []):
                name_text = cred.get("name", {}).get("nameText", {}).get("text")
                if name_text:
                    names.append(name_text)
                    
            if is_creator:
                creators_list.extend(names)
            elif is_star:
                stars_list.extend(names)
                
        creators_str = ", ".join(list(set(creators_list))) if creators_list else None
        stars_str = ", ".join(list(set(stars_list))) if stars_list else None
        
        # Parse genres
        genres = []
        genres_list = node.get("titleGenres", {}).get("genres", [])
        for g in genres_list:
            g_text = g.get("genre", {}).get("text")
            if g_text:
                genres.append(g_text)
                
        rank_obj = node.get("meterRanking", {})
        current_rank = rank_obj.get("currentRank") if rank_obj else (idx + 1)
        
        show_data = {
            "id": show_id,
            "title": title,
            "type": show_type,
            "release_year": release_year,
            "end_year": end_year,
            "global_rating": global_rating,
            "global_vote_count": global_vote_count,
            "runtime_seconds": runtime_seconds,
            "certificate": certificate,
            "plot": plot,
            "poster_url": poster_url,
            "release_date": release_date,
            "total_episodes": total_episodes,
            "creators": creators_str,
            "stars": stars_str,
            "genres": genres,
            "current_rank": current_rank
        }
        shows.append(show_data)
        
    return shows

def extract_country_ratings(next_data):
    """
    Extracts ratings by country from ratings page JSON.
    """
    country_ratings = []
    if not next_data:
        return country_ratings
        
    content_data = next_data.get("props", {}).get("pageProps", {}).get("contentData", {})
    
    # Option 1: props.pageProps.contentData.data.title.aggregateRatingsBreakdown.ratingsSummaryByCountry
    list1 = content_data.get("data", {}).get("title", {}).get("aggregateRatingsBreakdown", {}).get("ratingsSummaryByCountry", [])
    if list1:
        for item in list1:
            code = item.get("country")
            name = item.get("displayText", {}).get("value") if isinstance(item.get("displayText"), dict) else item.get("displayText")
            rating = item.get("aggregate")
            votes = item.get("voteCount")
            if code and name:
                country_ratings.append({
                    "country_code": code,
                    "country_name": name,
                    "rating": rating,
                    "vote_count": votes
                })
                
    # Option 2 Fallback: props.pageProps.contentData.histogramData.countryData
    if not country_ratings:
        list2 = content_data.get("histogramData", {}).get("countryData", [])
        for item in list2:
            code = item.get("countryCode")
            name = item.get("displayText")
            rating = item.get("aggregateRating")
            votes = item.get("totalVoteCount")
            if code and name:
                country_ratings.append({
                    "country_code": code,
                    "country_name": name,
                    "rating": rating,
                    "vote_count": votes
                })
                
    return country_ratings

def extract_reviews(next_data):
    """
    Extracts user reviews from reviews page JSON.
    """
    reviews = []
    if not next_data:
        return reviews
        
    page_props = next_data.get("props", {}).get("pageProps", {})
    content_data = page_props.get("contentData", {})
    
    # Option 1: props.pageProps.contentData.reviews
    list1 = content_data.get("reviews", [])
    for item in list1:
        r = item.get("review", {})
        if not r:
            continue
        
        author_info = r.get("author", {})
        author_username = author_info.get("displayName") or author_info.get("username", {}).get("text") or "Anonymous"
        author_id = author_info.get("userId") or ""
        
        h = r.get("helpfulnessVotes", {}) or {}
        up_votes = h.get("upVotes") or 0
        down_votes = h.get("downVotes") or 0
        
        r_summary = r.get("reviewSummary") or ""
        r_content = r.get("reviewText") or ""
        
        reviews.append({
            "id": r.get("reviewId"),
            "author_username": author_username,
            "author_id": author_id,
            "rating": r.get("authorRating"),
            "summary": r_summary,
            "content": r_content,
            "submission_date": r.get("submissionDate"),
            "up_votes": up_votes,
            "down_votes": down_votes,
            "is_spoiler": bool(r.get("spoiler"))
        })
        
    # Option 2 Fallback: props.pageProps.contentData.data.title.reviews.edges
    if not reviews:
        edges = content_data.get("data", {}).get("title", {}).get("reviews", {}).get("edges", [])
        for edge in edges:
            node = edge.get("node", {})
            if not node:
                continue
                
            author_info = node.get("author", {}) or {}
            author_username = author_info.get("username", {}).get("text") or "Anonymous"
            author_id = author_info.get("userId") or ""
            
            summary_obj = node.get("summary", {}) or {}
            r_summary = summary_obj.get("originalText") or summary_obj.get("text") or ""
            
            text_obj = node.get("text", {}) or {}
            r_content = text_obj.get("originalText") or text_obj.get("text") or ""
            
            h = node.get("helpfulness", {}) or {}
            up_votes = h.get("upVotes") or 0
            down_votes = h.get("downVotes") or 0
            
            reviews.append({
                "id": node.get("id"),
                "author_username": author_username,
                "author_id": author_id,
                "rating": node.get("authorRating"),
                "summary": r_summary,
                "content": r_content,
                "submission_date": node.get("submissionDate"),
                "up_votes": up_votes,
                "down_votes": down_votes,
                "is_spoiler": bool(node.get("spoiler"))
            })
            
    # Apply limit
    return reviews[:REVIEWS_LIMIT]

def validate_date(date_str):
    if not date_str:
        return None
    date_str = str(date_str).strip()
    if date_str.lower() in ["release date", "date", "null", "none", ""]:
        return None
    import datetime
    # Try parsing multiple date formats and normalize to YYYY-MM-DD
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            dt = datetime.datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Try to extract year-only format if the string is just 4 digits
    if date_str.isdigit() and len(date_str) == 4:
        return f"{date_str}-01-01"
    return None

def read_excel_shows(filepath):
    """
    Reads show data from Excel file. Returns (shows_list, platform_gender_list).
    Excel has headers at row 6: Week, MARKET, Platform, Content format, PAID/FREE, Content Type, Shows/Live Content, Language, Genre, Release Date, Rank, Reach
    Platform gender data starts at row 60.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]  # Use first sheet
    
    # Save metadata (week/time period and market) to metadata.json
    time_period = ws.cell(row=1, column=6).value
    market = ws.cell(row=7, column=2).value or "ALL INDIA"
    if time_period:
        try:
            week_val = ws.cell(row=7, column=1).value or "WK-26,2026"
            week_key = str(week_val).strip().replace(" ", "")
            
            existing_meta = {}
            if os.path.exists("metadata.json"):
                try:
                    with open("metadata.json", "r") as f:
                        loaded = json.load(f)
                        if isinstance(loaded, dict) and "time_period" not in loaded:
                            existing_meta = loaded
                        elif isinstance(loaded, dict) and "time_period" in loaded:
                            # Convert old single-week format to dict
                            old_key = "WK-26,2026"
                            if "WK-27" in loaded["time_period"]:
                                old_key = "WK-27,2026"
                            existing_meta = {old_key: loaded}
                except Exception:
                    pass
            
            # Clean up potential typos (e.g. if the cell text starts with WK-26 when it is actually WK-27)
            clean_time_period = str(time_period).strip()
            if "WK-27" in week_key and clean_time_period.startswith("WK-26"):
                clean_time_period = clean_time_period.replace("WK-26", "WK-27", 1)
            
            existing_meta[week_key] = {
                "time_period": clean_time_period,
                "market": str(market).strip()
            }
            with open("metadata.json", "w") as f:
                json.dump(existing_meta, f, indent=4)
            print(f"Saved metadata for {week_key}: {clean_time_period} | {market}")
        except Exception as e:
            print(f"Warning: Could not save metadata.json: {e}")
            
    col_map = {
        'week': 1,
        'market': 2,
        'platform': 3,
        'content_format': 4,
        'paid_free': 5,
        'content_type': 6,
        'title': 7,
        'language': 8,
        'genre': 9,
        'release_date': 10,
        'current_rank': 11,
        'reach': 12,
        'total_time_spent': None
    }
    
    # Try dynamically mapping columns based on row 6 headers
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=6, column=col).value
        if not header_val:
            continue
        h_str = str(header_val).strip().lower()
        if "week" in h_str:
            col_map['week'] = col
        elif "market" in h_str:
            col_map['market'] = col
        elif "platform" in h_str:
            col_map['platform'] = col
        elif "content format" in h_str:
            col_map['content_format'] = col
        elif "paid/free" in h_str or "paid_free" in h_str:
            col_map['paid_free'] = col
        elif "content type" in h_str:
            col_map['content_type'] = col
        elif "shows/live content" in h_str or h_str == "shows" or h_str == "title":
            col_map['title'] = col
        elif "language" in h_str:
            col_map['language'] = col
        elif "genre" in h_str:
            col_map['genre'] = col
        elif "release date" in h_str:
            col_map['release_date'] = col
        elif "rank" in h_str:
            col_map['current_rank'] = col
        elif "reach" in h_str:
            col_map['reach'] = col
        elif "time spent" in h_str or "total time spent" in h_str:
            col_map['total_time_spent'] = col

    shows = []
    # Data rows start at row 7
    for row in range(7, ws.max_row + 1):
        # Stop show loop if we reach the platform gender section header/content
        if row >= 60:
            val_col1 = str(ws.cell(row=row, column=1).value or "").strip().upper()
            if val_col1 in ["PLATFORM", "JIOHOTSTAR", "AMAZON PRIME VIDEO", "NETFLIX", "DISNEY+", "ZEE5", "SONYLIV"]:
                break

        title = ws.cell(row=row, column=col_map['title']).value  # Shows/Live Content
        if not title:
            continue
            
        title_str = str(title).strip()
        if title_str.lower() in ["shows/live content", "shows", "title", "shows/live content "]:
            continue
        
        release_date = ws.cell(row=row, column=col_map['release_date']).value
        if release_date and hasattr(release_date, 'strftime'):
            release_date = release_date.strftime('%Y-%m-%d')
        else:
            release_date = str(release_date).strip() if release_date else None
            
        release_date = validate_date(release_date)
        
        genres_str = ws.cell(row=row, column=col_map['genre']).value or ""
        genres = [g.strip() for g in genres_str.split(',') if g.strip()]
        
        total_time_spent = None
        if col_map['total_time_spent']:
            spent_val = ws.cell(row=row, column=col_map['total_time_spent']).value
            if spent_val is not None:
                try:
                    total_time_spent = float(spent_val)
                except (ValueError, TypeError):
                    pass
        
        show = {
            'title': str(title).strip(),
            'week': ws.cell(row=row, column=col_map['week']).value,
            'market': ws.cell(row=row, column=col_map['market']).value,
            'platform': ws.cell(row=row, column=col_map['platform']).value,
            'content_format': ws.cell(row=row, column=col_map['content_format']).value,
            'paid_free': ws.cell(row=row, column=col_map['paid_free']).value,
            'content_type': ws.cell(row=row, column=col_map['content_type']).value,
            'languages': ws.cell(row=row, column=col_map['language']).value,
            'genres': genres,
            'release_date': release_date,
            'current_rank': ws.cell(row=row, column=col_map['current_rank']).value,
            'reach': ws.cell(row=row, column=col_map['reach']).value,
            'total_time_spent': total_time_spent
        }
        shows.append(show)
    
    # Read platform gender data (starts around row 60)
    platform_gender = []
    for row in range(60, ws.max_row + 1):
        platform = ws.cell(row=row, column=1).value
        total_reach = ws.cell(row=row, column=2).value
        male_pct = ws.cell(row=row, column=3).value
        female_pct = ws.cell(row=row, column=4).value
        
        if platform and platform != 'PLATFORM' and total_reach is not None:
            platform_gender.append({
                'platform': platform,
                'total_reach': total_reach,
                'male_pct': male_pct,
                'female_pct': female_pct
            })
    
    return shows, platform_gender

def clean_show_title(title):
    import re
    cleaned = title
    # Remove things like "Season 1", "Season 02", "season 3", "S1", "S01", "s2"
    cleaned = re.sub(r'\b(season|series|vol|volume|v)\.?\s*\d+\b', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\bS\d{1,2}\b', '', cleaned, flags=re.IGNORECASE)
    # Remove year like (2024), [2024], 2024
    cleaned = re.sub(r'[\(\[\{]?\b\d{4}\b[\)\]\}]?', '', cleaned)
    # Remove trailing/leading colons, dashes, spaces
    cleaned = re.sub(r'[\s\-:\,]+$', '', cleaned)
    cleaned = re.sub(r'^[\s\-:\,]+', '', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned

def search_imdb_show(driver, original_title):
    """
    Searches IMDb for a show title and returns the first result's ID, or None.
    Uses query cleaning and scoring to filter the best matching TV Series or Movie.
    """
    import urllib.parse
    import re
    
    cleaned_title = clean_show_title(original_title)
    
    # Try searching with cleaned title first, fallback to original if they are different
    queries_to_try = [cleaned_title]
    if cleaned_title != original_title:
        queries_to_try.append(original_title)
        
    for query in queries_to_try:
        search_url = f"https://www.imdb.com/find/?q={urllib.parse.quote(query)}&s=tt"
        try:
            delay = random.uniform(1.0, 2.0)
            time.sleep(delay)
            driver.get(search_url)
            time.sleep(3.0)
            
            soup = BeautifulSoup(driver.page_source, "lxml")
            containers = soup.find_all(class_="ipc-metadata-list-summary-item")
            
            candidates = []
            for c in containers:
                links = c.find_all('a', href=lambda h: h and '/title/tt' in h)
                link = None
                for l in links:
                    if l.text.strip():
                        link = l
                        break
                if not link and links:
                    link = links[0]
                if not link:
                    continue
                    
                href = link.get('href', '')
                match = re.search(r'(tt\d+)', href)
                if not match:
                    continue
                show_id = match.group(1)
                
                title_text = link.text.strip()
                container_text = c.get_text(" ")
                
                # Check for poster image in the container
                has_poster = (c.select_one('.ipc-media__img') is not None) or (c.find('img') is not None)
                
                # Calculate match score
                score = 0
                
                t_lower = title_text.lower()
                q_lower = query.lower()
                
                if t_lower == q_lower:
                    score += 100
                elif q_lower in t_lower:
                    score += 50
                elif t_lower in q_lower:
                    score += 40
                else:
                    # Keyword matching
                    q_words = set(re.findall(r'\w+', q_lower))
                    t_words = set(re.findall(r'\w+', t_lower))
                    intersection = q_words.intersection(t_words)
                    score += len(intersection) * 15
                    
                # Poster bonus
                if has_poster:
                    score += 20
                    
                # Type penalty/bonus
                lower_ct = container_text.lower()
                if "tv episode" in lower_ct or "podcast episode" in lower_ct:
                    score -= 60
                elif "tv series" in lower_ct:
                    score += 30
                elif "movie" in lower_ct:
                    score += 20
                elif "podcast series" in lower_ct:
                    score -= 40
                    
                candidates.append({
                    'id': show_id,
                    'title': title_text,
                    'score': score,
                    'has_poster': has_poster
                })
                
            if candidates:
                candidates.sort(key=lambda x: x['score'], reverse=True)
                best = candidates[0]
                if best['score'] >= 40:
                    return best['id']
                    
        except (TimeoutException, WebDriverException) as e:
            raise e
        except Exception as e:
            print(f"  Error searching for query '{query}': {e}")
            
    return None


def main():
    print("IMDb Show Scraper & Importer Initialization...")
    
    # Create a lock file to indicate that the scraper is active
    with open("scraper.lock", "w") as f:
        f.write("running")
        
    # Make sure tables exist
    db.init_db()
    
    # Determine import path (argument directory or default)
    import_path = sys.argv[1] if len(sys.argv) > 1 else None
    
    # Proactively check data_input directory if no argument is passed
    if not import_path:
        if os.path.exists("data_input") and os.path.isdir("data_input"):
            import_path = "data_input"
            
    excel_files = []
    if import_path:
        if os.path.isdir(import_path):
            # It's a directory! Scan for all .xlsx files
            for f in os.listdir(import_path):
                if f.endswith(".xlsx") and not f.startswith("~$"):
                    excel_files.append(os.path.join(import_path, f))
        elif os.path.isfile(import_path) and import_path.endswith(".xlsx"):
            excel_files.append(import_path)
            
    # Fallback to default single Excel file if no directory files found
    if not excel_files and os.path.exists(EXCEL_FILE):
        excel_files.append(EXCEL_FILE)
        
    if excel_files:
        print(f"Found {len(excel_files)} Excel files to process.")
        
        # Load week codes to sort chronologically
        files_with_weeks = []
        for filepath in excel_files:
            try:
                wb = openpyxl.load_workbook(filepath, read_only=True)
                ws = wb[wb.sheetnames[0]]
                # Try to find week code in data cells first (row 7 column 1)
                week_code = ws.cell(row=7, column=1).value
                if not week_code:
                    # Try searching rows 7 to 15
                    for r in range(7, 15):
                        val = ws.cell(row=r, column=1).value
                        if val:
                            week_code = val
                            break
                # If still not found, try sheet title or metadata cell (1, 6)
                if not week_code:
                    week_code = ws.cell(row=1, column=6).value
                    
                files_with_weeks.append((filepath, week_code))
                wb.close()
            except Exception as e:
                print(f"Warning: Could not read week code from {filepath}: {e}")
                files_with_weeks.append((filepath, "WK-00, 2026"))
                
        # Sort by week code chronologically
        def parse_week_sort_key(item):
            filepath, week_code = item
            if not week_code:
                return (0, 0)
            import re
            # Extract WK-XX and Year
            match = re.search(r'WK-(\d+)\s*,\s*(\d+)', str(week_code))
            if match:
                return (int(match.group(2)), int(match.group(1)))
            return (0, 0)
            
        files_with_weeks.sort(key=parse_week_sort_key)
        sorted_files = [item[0] for item in files_with_weeks]
        
        print("Sorted Excel files chronologically:")
        for idx, f in enumerate(sorted_files, 1):
            print(f"  {idx}. {f} (Week: {files_with_weeks[idx-1][1]})")
            
        # Connect to DB
        conn, is_sqlite = db.get_connection()
        driver = None
        
        latest_time_period = None
        latest_market = "ALL INDIA"
        
        try:
            for filepath in sorted_files:
                print(f"\n=========================================")
                print(f"Processing File: {filepath}")
                print(f"=========================================")
                
                # Load metadata
                try:
                    wb_meta = openpyxl.load_workbook(filepath, data_only=True)
                    ws_meta = wb_meta[wb_meta.sheetnames[0]]
                    time_period = ws_meta.cell(row=1, column=6).value
                    market = ws_meta.cell(row=7, column=2).value or "ALL INDIA"
                    if time_period:
                        latest_time_period = str(time_period).strip()
                    if market:
                        latest_market = str(market).strip()
                    wb_meta.close()
                except Exception as meta_err:
                    print(f"Warning: Could not read metadata cell: {meta_err}")
                
                # Load shows from this file
                excel_shows, platform_gender = read_excel_shows(filepath)
                print(f"Loaded {len(excel_shows)} shows from {filepath}.")
                
                # Save platform gender
                try:
                    for pg in platform_gender:
                        db.save_platform_gender(conn, is_sqlite, pg)
                    conn.commit()
                    print(f"Saved platform gender records.")
                except Exception as e:
                    conn.rollback()
                    print(f"Warning: Failed to save platform gender: {e}")
                    
                # Loop through shows
                for i, excel_show in enumerate(excel_shows, 1):
                    title = excel_show['title']
                    week_code = excel_show.get('week') or "WK-26,2026"
                    
                    # 1. Check if show already exists in db by title
                    existing_id = None
                    cursor = conn.cursor()
                    query_check = "SELECT id FROM shows WHERE LOWER(TRIM(title)) = %s LIMIT 1"
                    if is_sqlite:
                        query_check = query_check.replace("%s", "?")
                    try:
                        cursor.execute(query_check, (title.lower().strip(),))
                        row = cursor.fetchone()
                        if row:
                            existing_id = row[0]
                    except Exception as e:
                        print(f"Error checking show existence: {e}")
                    
                    show_id = existing_id
                    
                    if show_id:
                        print(f"  [{i}/{len(excel_shows)}] Show '{title}' exists in DB (ID: {show_id}). Skipping IMDb scrape.")
                    else:
                        print(f"  [{i}/{len(excel_shows)}] Show '{title}' is NEW. Searching on IMDb...")
                        
                        # Start driver on-demand (lazy load)
                        if not driver:
                            print("  Lazy loading headless Chrome webdriver...")
                            driver = init_driver()
                            
                        # Search IMDb
                        show_id = search_imdb_show(driver, title)
                        
                        if show_id:
                            print(f"    Found IMDb ID: {show_id}")
                            # Scrape details
                            detail_url = f"https://www.imdb.com/title/{show_id}/"
                            detail_json = get_next_data(driver, detail_url)
                            
                            imdb_data = {}
                            imdb_genres = []
                            if detail_json:
                                page_props = detail_json.get('props', {}).get('pageProps', {})
                                above_fold = page_props.get('aboveTheFoldData', {})
                                if above_fold:
                                    ratings = above_fold.get('ratingsSummary', {})
                                    imdb_data['global_rating'] = ratings.get('aggregateRating')
                                    imdb_data['global_vote_count'] = ratings.get('voteCount')
                                    
                                    cert = above_fold.get('certificate', {})
                                    imdb_data['certificate'] = cert.get('rating') if cert else None
                                    
                                    runtime_obj = above_fold.get('runtime', {})
                                    imdb_data['runtime_seconds'] = runtime_obj.get('seconds') if runtime_obj else None
                                    
                                    release_year_obj = above_fold.get('releaseYear', {})
                                    imdb_data['release_year'] = release_year_obj.get('year') if release_year_obj else None
                                    imdb_data['end_year'] = release_year_obj.get('endYear') if release_year_obj else None
                                    
                                    title_type = above_fold.get('titleType', {})
                                    imdb_data['type'] = title_type.get('id') if title_type else None
                                    
                                    plot_obj = above_fold.get('plot', {})
                                    if plot_obj:
                                        plot_text = plot_obj.get('plotText', {})
                                        imdb_data['plot'] = plot_text.get('plainText') if plot_text else None
                                    
                                    img = above_fold.get('primaryImage', {})
                                    imdb_data['poster_url'] = img.get('url') if img else None
                                    if not imdb_data['poster_url'] and detail_json.get('fallback_poster_url'):
                                        imdb_data['poster_url'] = detail_json['fallback_poster_url']
                                    
                                    # Parse genres
                                    genres_obj = above_fold.get('genres', {}).get('genres', [])
                                    for g in genres_obj:
                                        g_text = g.get('text')
                                        if g_text:
                                            imdb_genres.append(g_text)
                                            
                                    # Parse creators and stars
                                    # IMDb uses principalCreditsV2 in aboveTheFoldData (grouping.text)
                                    creators_list = []
                                    stars_list = []

                                    for credit in above_fold.get('principalCreditsV2', []):
                                        grouping_text = credit.get('grouping', {}).get('text', '').lower()
                                        names = [c.get('name', {}).get('nameText', {}).get('text', '') for c in credit.get('credits', [])]
                                        names = [n for n in names if n]
                                        if 'creator' in grouping_text or 'director' in grouping_text or 'writer' in grouping_text:
                                            creators_list.extend(names)
                                        elif 'star' in grouping_text or 'cast' in grouping_text or 'actor' in grouping_text:
                                            stars_list.extend(names)

                                    # Fallback: mainColumnData principalCredits (category.text)
                                    if not creators_list and not stars_list:
                                        main_col_data = page_props.get('mainColumnData', {})
                                        for credit in main_col_data.get('principalCredits', []):
                                            category = credit.get('category', {}).get('text', '').lower()
                                            names = [c.get('name', {}).get('nameText', {}).get('text', '') for c in credit.get('credits', [])]
                                            names = [n for n in names if n]
                                            if 'creator' in category or 'director' in category or 'writer' in category:
                                                creators_list.extend(names)
                                            elif 'star' in category or 'cast' in category or 'actor' in category:
                                                stars_list.extend(names)

                                    imdb_data['creators'] = ', '.join(creators_list) if creators_list else None
                                    imdb_data['stars'] = ', '.join(stars_list) if stars_list else None
                                    

                            # Fetch country ratings & reviews
                            ratings_url = f"https://www.imdb.com/title/{show_id}/ratings/"
                            ratings_json = get_next_data(driver, ratings_url)
                            country_ratings = extract_country_ratings(ratings_json)
                            
                            reviews_url = f"https://www.imdb.com/title/{show_id}/reviews/"
                            try: driver.delete_all_cookies()
                            except Exception: pass
                            reviews_json = get_next_data(driver, reviews_url)
                            reviews = extract_reviews(reviews_json)
                            
                            # Save static details
                            show_data = {
                                'id': show_id,
                                'title': title,
                                'type': imdb_data.get('type', excel_show.get('content_type')),
                                'release_year': imdb_data.get('release_year'),
                                'end_year': imdb_data.get('end_year'),
                                'global_rating': imdb_data.get('global_rating'),
                                'global_vote_count': imdb_data.get('global_vote_count'),
                                'runtime_seconds': imdb_data.get('runtime_seconds'),
                                'certificate': imdb_data.get('certificate'),
                                'plot': imdb_data.get('plot'),
                                'poster_url': imdb_data.get('poster_url'),
                                'release_date': excel_show.get('release_date'),
                                'total_episodes': None,
                                'creators': imdb_data.get('creators'),
                                'stars': imdb_data.get('stars'),
                                'current_rank': excel_show.get('current_rank'),
                                'platform': excel_show.get('platform'),
                                'content_format': excel_show.get('content_format'),
                                'paid_free': excel_show.get('paid_free'),
                                'content_type': excel_show.get('content_type'),
                                'languages': excel_show.get('languages'),
                                'reach': excel_show.get('reach'),
                                'week': excel_show.get('week'),
                                'market': excel_show.get('market'),
                            }
                            try:
                                db.save_show(conn, is_sqlite, show_data)
                                all_genres = list(set(excel_show.get('genres', []) + imdb_genres))
                                db.save_genres(conn, is_sqlite, show_id, all_genres)
                                db.save_country_ratings(conn, is_sqlite, show_id, country_ratings)
                                db.save_reviews(conn, is_sqlite, show_id, reviews)
                                conn.commit()
                                print(f"    Successfully scraped and saved '{title}' details to DB.")
                            except Exception as save_err:
                                conn.rollback()
                                print(f"    Error saving scraped data for '{title}': {save_err}")
                        else:
                            print(f"    IMDb search failed for '{title}'. Saving with Excel details only.")
                            import hashlib
                            show_id = "xl_" + hashlib.md5(title.lower().strip().encode('utf-8')).hexdigest()[:10]
                            show_data = {
                                'id': show_id,
                                'title': title,
                                'type': excel_show.get('content_type'),
                                'release_year': None,
                                'end_year': None,
                                'global_rating': None,
                                'global_vote_count': None,
                                'runtime_seconds': None,
                                'certificate': None,
                                'plot': None,
                                'poster_url': None,
                                'release_date': excel_show.get('release_date'),
                                'total_episodes': None,
                                'creators': None,
                                'stars': None,
                                'current_rank': excel_show.get('current_rank'),
                                'platform': excel_show.get('platform'),
                                'content_format': excel_show.get('content_format'),
                                'paid_free': excel_show.get('paid_free'),
                                'content_type': excel_show.get('content_type'),
                                'languages': excel_show.get('languages'),
                                'reach': excel_show.get('reach'),
                                'week': excel_show.get('week'),
                                'market': excel_show.get('market'),
                            }
                            try:
                                db.save_show(conn, is_sqlite, show_data)
                                db.save_genres(conn, is_sqlite, show_id, excel_show.get('genres', []))
                                conn.commit()
                                print(f"    Saved '{title}' (unmapped) to DB.")
                            except Exception as save_err:
                                conn.rollback()
                                print(f"    Error saving unmapped show '{title}': {save_err}")
                                
                    # 2. Save the weekly ranking
                    weekly_ranking = {
                        'show_id': show_id,
                        'week': week_code,
                        'current_rank': excel_show.get('current_rank'),
                        'reach': excel_show.get('reach'),
                        'platform': excel_show.get('platform'),
                        'content_format': excel_show.get('content_format'),
                        'paid_free': excel_show.get('paid_free'),
                        'content_type': excel_show.get('content_type'),
                        'market': excel_show.get('market')
                    }
                    try:
                        db.save_weekly_ranking(conn, is_sqlite, weekly_ranking)
                        
                        # Also update shows table fields with the latest rank/reach to preserve backward compatibility
                        update_query = """
                        UPDATE shows SET 
                            current_rank = %s, reach = %s, week = %s, platform = %s,
                            content_format = %s, paid_free = %s, content_type = %s, market = %s
                        WHERE id = %s
                        """
                        if is_sqlite:
                            update_query = update_query.replace("%s", "?")
                        cursor = conn.cursor()
                        cursor.execute(update_query, (
                            excel_show.get('current_rank'),
                            excel_show.get('reach'),
                            week_code,
                            db.normalize_platform(excel_show.get('platform')) if excel_show.get('platform') else None,
                            db.normalize_content_format(excel_show.get('content_format')) if excel_show.get('content_format') else None,
                            db.normalize_paid_free(excel_show.get('paid_free')) if excel_show.get('paid_free') else None,
                            db.normalize_content_type(excel_show.get('content_type')) if excel_show.get('content_type') else None,
                            db.normalize_market(excel_show.get('market')) if excel_show.get('market') else "ALL INDIA",
                            show_id
                        ))
                        conn.commit()
                    except Exception as wr_err:
                        conn.rollback()
                        print(f"  Error saving weekly ranking: {wr_err}")
            
            print("\nExcel Import process finished successfully.")
        finally:
            conn.close()
            if driver:
                driver.quit()
                print("Headless browser closed.")
    else:
        # TV METER MODE (original behavior)
        chart_url = "https://www.imdb.com/chart/tvmeter/"
        print("Fetching trending TV shows from IMDb TV Meter...")
        driver = init_driver()
        try:
            chart_json = get_next_data(driver, chart_url)
            if not chart_json:
                print("Error: Could not retrieve TV Meter data. Aborting.")
                return
                
            shows = extract_tv_shows_list(chart_json)
            print(f"Extracted {len(shows)} shows from chart list.")
            if not shows:
                print("Error: No shows could be parsed. Aborting.")
                return
                
            shows_to_scrape = shows[:SCRAPE_LIMIT]
            print(f"Configured to scrape details for {len(shows_to_scrape)} shows (SCRAPE_LIMIT = {SCRAPE_LIMIT}).")
            
            conn, is_sqlite = db.get_connection()
            try:
                db.clear_all_ranks(conn, is_sqlite)
                for show in shows_to_scrape:
                    db.save_show(conn, is_sqlite, show)
                conn.commit()
                print(f"  Updated ranks for {len(shows_to_scrape)} shows.")
            except Exception as rank_err:
                conn.rollback()
                print(f"Warning: Could not update rankings: {rank_err}")
            
            for i, show in enumerate(shows_to_scrape, 1):
                show_id = show['id']
                title = show['title']
                
                if db.is_show_fresh(conn, is_sqlite, show_id):
                    print(f"\n--- [{i}/{len(shows_to_scrape)}] Skipping {title} ({show_id}) - Data is already fresh ---")
                    try:
                        db.save_show(conn, is_sqlite, show)
                        conn.commit()
                    except Exception as e:
                        conn.rollback()
                    continue
                    
                print(f"\n--- [{i}/{len(shows_to_scrape)}] Processing show: {title} ({show_id}) ---")
                
                ratings_url = f"https://www.imdb.com/title/{show_id}/ratings/"
                ratings_json = get_next_data(driver, ratings_url)
                country_ratings = extract_country_ratings(ratings_json)
                print(f"  Found {len(country_ratings)} country-specific rating records.")
                
                reviews_url = f"https://www.imdb.com/title/{show_id}/reviews/"
                try:
                    driver.delete_all_cookies()
                except Exception:
                    pass
                reviews_json = get_next_data(driver, reviews_url)
                reviews = extract_reviews(reviews_json)
                
                try:
                    db.save_show(conn, is_sqlite, show)
                    db.save_genres(conn, is_sqlite, show_id, show['genres'])
                    db.save_country_ratings(conn, is_sqlite, show_id, country_ratings)
                    db.save_reviews(conn, is_sqlite, show_id, reviews)
                    conn.commit()
                    print(f"  Successfully saved {title} data to the database.")
                except Exception as e:
                    conn.rollback()
                    print(f"  Error: Failed to save {title} data to DB: {e}")
            print("\nIMDb TV Meter Scraping process finished successfully.")
        finally:
            driver.quit()
            conn.close()
            print("Database connection & browser closed.")

    if os.path.exists("scraper.lock"):
        os.remove("scraper.lock")

def update_job_progress(conn, is_sqlite, job_id, status, processed_shows, total_shows, current_show=None, error_message=None):
    if job_id is None:
        return
    query = """
    UPDATE scraping_jobs 
    SET status = %s, processed_shows = %s, total_shows = %s, current_show = %s, error_message = %s, updated_at = CURRENT_TIMESTAMP
    WHERE id = %s
    """
    if is_sqlite:
        query = query.replace("%s", "?")
    try:
        cursor = conn.cursor()
        cursor.execute(query, (status, processed_shows, total_shows, current_show, error_message, job_id))
        conn.commit()
    except Exception as e:
        print(f"Error updating job progress: {e}")

def check_job_status(conn, is_sqlite, job_id):
    if not job_id:
        return "running"
    try:
        conn.commit()
    except Exception:
        pass
    
    try:
        cursor = conn.cursor()
        query = "SELECT status FROM scraping_jobs WHERE id = %s"
        if is_sqlite:
            query = query.replace("%s", "?")
        cursor.execute(query, (job_id,))
        row = cursor.fetchone()
        cursor.close()
        if row:
            return row[0]
    except Exception as e:
        print(f"Error checking job status in thread: {e}")
    return "running"

def process_excel_file(filepath, job_id=None):
    """
    Processes an uploaded Excel file, scrapes IMDb details in the background, 
    and writes progress states to the database.
    """
    conn, is_sqlite = db.get_connection()
    driver = None
    
    try:
        # Load metadata
        wb_meta = openpyxl.load_workbook(filepath, data_only=True)
        ws_meta = wb_meta[wb_meta.sheetnames[0]]
        time_period = ws_meta.cell(row=1, column=6).value
        market = ws_meta.cell(row=7, column=2).value or "ALL INDIA"
        
        # Read week code
        week_code = ws_meta.cell(row=7, column=1).value
        if not week_code:
            for r in range(7, 15):
                val = ws_meta.cell(row=r, column=1).value
                if val:
                    week_code = val
                    break
        if not week_code:
            week_code = time_period
        if not week_code:
            week_code = "WK-26, 2026"
            
        wb_meta.close()
        
        # Load shows from this file
        excel_shows, platform_gender = read_excel_shows(filepath)
        total_shows = len(excel_shows)
        
        # Update job to running
        update_job_progress(conn, is_sqlite, job_id, "running", 0, total_shows, "Started reading Excel file")
        
        # Save platform gender
        for pg in platform_gender:
            db.save_platform_gender(conn, is_sqlite, pg)
        conn.commit()
        
        # Loop through shows
        for i, excel_show in enumerate(excel_shows, 1):
            title = excel_show['title']
            show_week = excel_show.get('week') or week_code
            
            # Check job status for Pause / Cancel
            job_status = check_job_status(conn, is_sqlite, job_id)
            while job_status == "paused":
                time.sleep(2.0)
                job_status = check_job_status(conn, is_sqlite, job_id)
                
            if job_status in ["cancelled", "stopped"]:
                print(f"Job {job_id} cancelled/stopped by user. Exiting thread.")
                if driver:
                    try: driver.quit()
                    except Exception: pass
                return
            
            # Update job progress
            update_job_progress(conn, is_sqlite, job_id, "running", i - 1, total_shows, f"Scraping '{title}' ({i}/{total_shows})")
            
            # Check if show already exists in db by title
            existing_id = None
            cursor = conn.cursor()
            query_check = "SELECT id FROM shows WHERE LOWER(TRIM(title)) = %s LIMIT 1"
            if is_sqlite:
                query_check = query_check.replace("%s", "?")
            
            cursor.execute(query_check, (title.lower().strip(),))
            row = cursor.fetchone()
            cursor.close()
            if row:
                existing_id = row[0]
                
            show_id = existing_id
            
            if show_id:
                # Existing show - skip scrape
                print(f"Show {title} exists. Skipping scrape.")
            else:
                try:
                    # Scrape on-demand
                    if not driver:
                        driver = init_driver()
                    
                    show_id = search_imdb_show(driver, title)
                    if show_id:
                        # Scrape details
                        detail_url = f"https://www.imdb.com/title/{show_id}/"
                        detail_json = get_next_data(driver, detail_url)
                        
                        imdb_data = {}
                        imdb_genres = []
                        if detail_json:
                            page_props = detail_json.get('props', {}).get('pageProps', {})
                            above_fold = page_props.get('aboveTheFoldData', {})
                            if above_fold:
                                ratings = above_fold.get('ratingsSummary', {})
                                imdb_data['global_rating'] = ratings.get('aggregateRating')
                                imdb_data['global_vote_count'] = ratings.get('voteCount')
                                
                                cert = above_fold.get('certificate', {})
                                imdb_data['certificate'] = cert.get('rating') if cert else None
                                
                                runtime_obj = above_fold.get('runtime', {})
                                imdb_data['runtime_seconds'] = runtime_obj.get('seconds') if runtime_obj else None
                                
                                release_year_obj = above_fold.get('releaseYear', {})
                                imdb_data['release_year'] = release_year_obj.get('year') if release_year_obj else None
                                imdb_data['end_year'] = release_year_obj.get('endYear') if release_year_obj else None
                                
                                title_type = above_fold.get('titleType', {})
                                imdb_data['type'] = title_type.get('id') if title_type else None
                                
                                plot_obj = above_fold.get('plot', {})
                                if plot_obj:
                                    plot_text = plot_obj.get('plotText', {})
                                    imdb_data['plot'] = plot_text.get('plainText') if plot_text else None
                                
                                img = above_fold.get('primaryImage', {})
                                imdb_data['poster_url'] = img.get('url') if img else None
                                if not imdb_data['poster_url'] and detail_json.get('fallback_poster_url'):
                                    imdb_data['poster_url'] = detail_json['fallback_poster_url']
                                
                                # Parse genres
                                genres_obj = above_fold.get('genres', {}).get('genres', [])
                                for g in genres_obj:
                                    g_text = g.get('text')
                                    if g_text:
                                        imdb_genres.append(g_text)
                                        
                                creators_list = []
                                stars_list = []
                                for credit in above_fold.get('principalCreditsV2', []):
                                    grouping_text = credit.get('grouping', {}).get('text', '').lower()
                                    names = [c.get('name', {}).get('nameText', {}).get('text', '') for c in credit.get('credits', [])]
                                    names = [n for n in names if n]
                                    if 'creator' in grouping_text or 'director' in grouping_text or 'writer' in grouping_text:
                                        creators_list.extend(names)
                                    elif 'star' in grouping_text or 'cast' in grouping_text or 'actor' in grouping_text:
                                        stars_list.extend(names)
                                
                                if not creators_list and not stars_list:
                                    main_col_data = page_props.get('mainColumnData', {})
                                    for credit in main_col_data.get('principalCredits', []):
                                        category = credit.get('category', {}).get('text', '').lower()
                                        names = [c.get('name', {}).get('nameText', {}).get('text', '') for c in credit.get('credits', [])]
                                        names = [n for n in names if n]
                                        if 'creator' in category or 'director' in category or 'writer' in category:
                                            creators_list.extend(names)
                                        elif 'star' in category or 'cast' in category or 'actor' in category:
                                            stars_list.extend(names)
                                
                                imdb_data['creators'] = ', '.join(creators_list) if creators_list else None
                                imdb_data['stars'] = ', '.join(stars_list) if stars_list else None
                                
                        ratings_url = f"https://www.imdb.com/title/{show_id}/ratings/"
                        ratings_json = get_next_data(driver, ratings_url)
                        country_ratings = extract_country_ratings(ratings_json)
                        
                        reviews_url = f"https://www.imdb.com/title/{show_id}/reviews/"
                        try: driver.delete_all_cookies()
                        except Exception: pass
                        reviews_json = get_next_data(driver, reviews_url)
                        reviews = extract_reviews(reviews_json)
                        
                        show_data = {
                            'id': show_id,
                            'title': title,
                            'type': imdb_data.get('type', excel_show.get('content_type')),
                            'release_year': imdb_data.get('release_year'),
                            'end_year': imdb_data.get('end_year'),
                            'global_rating': imdb_data.get('global_rating'),
                            'global_vote_count': imdb_data.get('global_vote_count'),
                            'runtime_seconds': imdb_data.get('runtime_seconds'),
                            'certificate': imdb_data.get('certificate'),
                            'plot': imdb_data.get('plot'),
                            'poster_url': imdb_data.get('poster_url'),
                            'release_date': excel_show.get('release_date'),
                            'total_episodes': None,
                            'creators': imdb_data.get('creators'),
                            'stars': imdb_data.get('stars'),
                            'current_rank': excel_show.get('current_rank'),
                            'platform': excel_show.get('platform'),
                            'content_format': excel_show.get('content_format'),
                            'paid_free': excel_show.get('paid_free'),
                            'content_type': excel_show.get('content_type'),
                            'languages': excel_show.get('languages'),
                            'reach': excel_show.get('reach'),
                            'week': excel_show.get('week'),
                            'market': excel_show.get('market'),
                            'total_time_spent': excel_show.get('total_time_spent')
                        }
                        db.save_show(conn, is_sqlite, show_data)
                        all_genres = list(set(excel_show.get('genres', []) + imdb_genres))
                        db.save_genres(conn, is_sqlite, show_id, all_genres)
                        db.save_country_ratings(conn, is_sqlite, show_id, country_ratings)
                        db.save_reviews(conn, is_sqlite, show_id, reviews)
                        conn.commit()
                    else:
                        import hashlib
                        show_id = "xl_" + hashlib.md5(title.lower().strip().encode('utf-8')).hexdigest()[:10]
                        show_data = {
                            'id': show_id,
                            'title': title,
                            'type': excel_show.get('content_type'),
                            'release_year': None,
                            'end_year': None,
                            'global_rating': None,
                            'global_vote_count': None,
                            'runtime_seconds': None,
                            'certificate': None,
                            'plot': None,
                            'poster_url': None,
                            'release_date': excel_show.get('release_date'),
                            'total_episodes': None,
                            'creators': None,
                            'stars': None,
                            'current_rank': excel_show.get('current_rank'),
                            'platform': excel_show.get('platform'),
                            'content_format': excel_show.get('content_format'),
                            'paid_free': excel_show.get('paid_free'),
                            'content_type': excel_show.get('content_type'),
                            'languages': excel_show.get('languages'),
                            'reach': excel_show.get('reach'),
                            'week': excel_show.get('week'),
                            'market': excel_show.get('market'),
                            'total_time_spent': excel_show.get('total_time_spent')
                        }
                        db.save_show(conn, is_sqlite, show_data)
                        db.save_genres(conn, is_sqlite, show_id, excel_show.get('genres', []))
                        conn.commit()
                except (TimeoutException, WebDriverException) as selenium_err:
                    print(f"Selenium error during scraping '{title}': {selenium_err}. Restarting driver...")
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    driver = None
                    
                    import hashlib
                    show_id = "xl_" + hashlib.md5(title.lower().strip().encode('utf-8')).hexdigest()[:10]
                    show_data = {
                        'id': show_id,
                        'title': title,
                        'type': excel_show.get('content_type'),
                        'release_year': None,
                        'end_year': None,
                        'global_rating': None,
                        'global_vote_count': None,
                        'runtime_seconds': None,
                        'certificate': None,
                        'plot': None,
                        'poster_url': None,
                        'release_date': excel_show.get('release_date'),
                        'total_episodes': None,
                        'creators': None,
                        'stars': None,
                        'current_rank': excel_show.get('current_rank'),
                        'platform': excel_show.get('platform'),
                        'content_format': excel_show.get('content_format'),
                        'paid_free': excel_show.get('paid_free'),
                        'content_type': excel_show.get('content_type'),
                        'languages': excel_show.get('languages'),
                        'reach': excel_show.get('reach'),
                        'week': excel_show.get('week'),
                        'market': excel_show.get('market'),
                        'total_time_spent': excel_show.get('total_time_spent')
                    }
                    db.save_show(conn, is_sqlite, show_data)
                    db.save_genres(conn, is_sqlite, show_id, excel_show.get('genres', []))
                    conn.commit()
                    time.sleep(5)
            
            # Save the weekly ranking
            weekly_ranking = {
                'show_id': show_id,
                'week': show_week,
                'current_rank': excel_show.get('current_rank'),
                'reach': excel_show.get('reach'),
                'platform': excel_show.get('platform'),
                'content_format': excel_show.get('content_format'),
                'paid_free': excel_show.get('paid_free'),
                'content_type': excel_show.get('content_type'),
                'market': excel_show.get('market'),
                'total_time_spent': excel_show.get('total_time_spent')
            }
            db.save_weekly_ranking(conn, is_sqlite, weekly_ranking)
            
            update_query = """
            UPDATE shows SET 
                current_rank = %s, reach = %s, week = %s, platform = %s,
                content_format = %s, paid_free = %s, content_type = %s, market = %s,
                total_time_spent = %s
            WHERE id = %s
            """
            if is_sqlite:
                update_query = update_query.replace("%s", "?")
            cursor = conn.cursor()
            cursor.execute(update_query, (
                excel_show.get('current_rank'),
                excel_show.get('reach'),
                show_week,
                db.normalize_platform(excel_show.get('platform')) if excel_show.get('platform') else None,
                db.normalize_content_format(excel_show.get('content_format')) if excel_show.get('content_format') else None,
                db.normalize_paid_free(excel_show.get('paid_free')) if excel_show.get('paid_free') else None,
                db.normalize_content_type(excel_show.get('content_type')) if excel_show.get('content_type') else None,
                db.normalize_market(excel_show.get('market')) if excel_show.get('market') else "ALL INDIA",
                excel_show.get('total_time_spent'),
                show_id
            ))
            cursor.close()
            conn.commit()
            
        # Update job to completed
        update_job_progress(conn, is_sqlite, job_id, "completed", total_shows, total_shows, "Completed scraping successfully")
        
    except Exception as e:
        print(f"Error in process_excel_file: {e}")
        import traceback
        traceback.print_exc()
        update_job_progress(conn, is_sqlite, job_id, "failed", 0, 0, error_message=str(e))
    finally:
        if driver:
            try: driver.quit()
            except Exception: pass
        conn.close()
        
        # Clean up uploaded excel file to free server storage
        if filepath and os.path.exists(filepath):
            try:
                os.remove(filepath)
                print(f"Cleaned up uploaded file: {filepath}")
            except Exception as remove_err:
                print(f"Warning: Could not remove uploaded file {filepath}: {remove_err}")

if __name__ == "__main__":
    main()
