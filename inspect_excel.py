import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\saxen\Downloads\imdb analysis.xlsx')
ws = wb['Top 50 Shows & Movies']

# Check rows 1-5 for title info
print('=== Rows 1-5 (all cells) ===')
for row in range(1, 6):
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            print(f'  ({row},{col}): {val}')

# Check all rows from 45 to end for platform data
print()
print('=== Rows 45 to end ===')
for row in range(45, ws.max_row + 1):
    vals = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            vals.append(f'Col{col}={v}')
    if vals:
        sep = '  |  '
        print(f'Row {row}: {sep.join(vals)}')
    else:
        print(f'Row {row}: (empty)')

# Count data rows (rows 7+)
print()
data_count = 0
for row in range(7, ws.max_row + 1):
    title = ws.cell(row=row, column=7).value  # Col 7 = Shows/Live Content
    if title is not None:
        data_count += 1
print(f'Total shows with data: {data_count}')
